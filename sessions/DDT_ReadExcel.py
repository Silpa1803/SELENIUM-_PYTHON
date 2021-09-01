import openpyxl

path="C:\Selenium Practice\DDT_ReadExcel.xlsx"

workbook=openpyxl.load_workbook(path)

sheet=workbook.active     #if you have sible worksheet in Excel
#sheet=workbook.get_sheet_by_name("DATA")  #if you have multiple sheet in Excel

rows=sheet.max_row      #13
cols=sheet.max_column    #4

print(rows)
print(cols)

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value,end='    ')
    print()
